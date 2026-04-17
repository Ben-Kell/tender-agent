from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Optional
import re

from openpyxl import load_workbook


RATE_CARD_PATH = Path("pricing/fujitsu_ictpa_pricing.xlsx")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _parse_money(value: object) -> Optional[float]:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = text.replace("$", "").replace(",", "").strip()

    try:
        return float(text)
    except ValueError:
        return None


def _extract_code_from_skill(skill_text: str) -> str:
    """
    Example:
    'Enterprise IT Governance GOVN' -> 'GOVN'
    'Strategic Planning ITSP' -> 'ITSP'
    """
    if not skill_text:
        return ""

    match = re.search(r"\b([A-Z]{3,5})\b$", skill_text.strip())
    if match:
        return match.group(1)

    all_codes = re.findall(r"\b([A-Z]{3,5})\b", skill_text)
    return all_codes[-1] if all_codes else ""


def _extract_skill_name(skill_text: str, code: str) -> str:
    if not skill_text:
        return ""

    if code and skill_text.endswith(code):
        return skill_text[: -len(code)].strip(" -")
    return skill_text.strip()


def load_ictpa_rate_card(path: str | Path = RATE_CARD_PATH) -> List[Dict[str, Any]]:
    """
    Reads the Fujitsu ICTPA pricing workbook and returns normalised rows.

    Expected source columns:
    C = Skills
    D = Level
    F = Tenderer's (Fujitsu) Labour Rate
    Optional context columns:
    A = Category
    B = Sub Category
    """
    workbook = load_workbook(filename=path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows: List[Dict[str, Any]] = []

    current_category = ""
    current_sub_category = ""

    # Start from Excel row 15 based on your screenshot / description.
    for excel_row in range(15, sheet.max_row + 1):
        category = _clean_text(sheet[f"A{excel_row}"].value)
        sub_category = _clean_text(sheet[f"B{excel_row}"].value)
        skill_text = _clean_text(sheet[f"C{excel_row}"].value)
        level_raw = sheet[f"D{excel_row}"].value
        rate_raw = sheet[f"F{excel_row}"].value

        if category:
            current_category = category
        if sub_category:
            current_sub_category = sub_category

        if not skill_text:
            continue

        try:
            level = int(level_raw) if level_raw not in (None, "") else None
        except Exception:
            level = None

        sell_rate = _parse_money(rate_raw)

        code = _extract_code_from_skill(skill_text)
        skill_name = _extract_skill_name(skill_text, code)

        rows.append({
            "source_row_number": excel_row,
            "category": current_category,
            "sub_category": current_sub_category,
            "skill_text": skill_text,
            "sfia_code": code,
            "skill_name": skill_name,
            "sfia_level": level,
            "unit": "day",
            "sell_rate": sell_rate,
        })

    return rows